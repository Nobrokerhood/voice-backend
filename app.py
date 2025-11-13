from flask import Flask, render_template, request, jsonify, send_file # <-- NEW: Import send_file
import requests
import io
import os
from flask_cors import CORS
import json
import google.generativeai as genai
import openpyxl # <-- NEW: Import openpyxl
from openpyxl.utils import get_column_letter # <-- NEW: Import helper
import logging

# --- Logging Configuration ---
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[logging.StreamHandler()])
# --- End of Logging Configuration ---

# --- Configuration ---
### 2. READ SECRETS FROM THE ENVIRONMENT ###
APPS_SCRIPT_BASE_URL = os.environ.get("APPS_SCRIPT_BASE_URL")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")

# --- End of Configuration ---

# --- Setup URLs ---
APPS_SCRIPT_URL_AUDIO = f"{APPS_SCRIPT_BASE_URL}?action=saveAudio"
APPS_SCRIPT_URL_LOG = f"{APPS_SCRIPT_BASE_URL}?action=logData"
# --- Define Template Schemas ---
TEMPLATE_SCHEMAS = {
    "template1": {
        "name": "Import Receipts",
        "schema": '''{
            "Payment Type": "e.g., Bank, Cheque, Cash",
            "Society Bank Name/Bank code": "The bank code, e.g., HDFC",
            "Cheque/Ref No": "The Cheque or Reference Number",
            "Tower No": "The tower number, e.g., 'B'",
            "Flat No": "The flat number, e.g., '502'",
            "Bill Head": "e.g., MAINTENANCE",
            "Amount": "The transaction amount as a number",
            "Transaction Date": "YYYY-MM-DD",
            "Comments": "Any user comments",
            "Meter No": "e.g., 6917",
            "Cheque Issuer Bank": "The name of the cheque issuer bank",
            "Cheque Date": "YYYY-MM-DD (if applicable)"
        }'''
    },
    "template2": {
        "name": "Vendor Bill Upload",
        "schema": '''{
            "Bill Number": "The vendor's bill number",
            "Bill Date": "YYYY-MM-DD",
            "Vendor Code": "The vendor's code",
            "Due Date": "YYYY-MM-DD",
            "Narration": "The description of the bill",
            "CGST Amount": "CGST amount as a number (default 0)",
            "SGST Amount": "SGST amount as a number (default 0)",
            "IGST Amount": "IGST amount as a number (default 0)",
            "TDS Amount": "TDS amount as a number (default 0)",
            "expenses": [
                {
                    "expense_code": "The code for the expense, e.g., 'ELEC_REPAIR'",
                    "expense_amount": "The amount for this specific expense as a number"
                }
            ]
        }'''
    },
    "template3": {
        "name": "Default / Not Specified",
        "schema": '''{
            "description": "A description of the entry",
            "amount": "The amount as a number"
        }'''
    }
}

# --- Configure Gemini ---
genai.configure(api_key=GEMINI_API_KEY)
gemini_model = genai.GenerativeModel('gemini-2.5-flash-lite')

app = Flask(__name__)
CORS(app)

@app.route('/')
def index():
  return render_template('index.html')

@app.route('/api/process-audio', methods=['POST'])
def process_audio():
  logging.info("Starting audio processing...")
  try:
    template_name = request.form['template']
    audio_file = request.files['audio']
    logging.info(f"Received request with template: {template_name} and audio file: {audio_file.filename}")

    audio_filename = audio_file.filename

    # --- 1. Store Audio in Google Drive ---
    logging.info(f"Uploading {audio_filename} to Google Drive...")
    audio_file.seek(0) # Ensure file pointer is at the beginning
    params = {'filename': audio_filename}
    headers = {'Content-Type': audio_file.mimetype}

    r_audio = requests.post(APPS_SCRIPT_URL_AUDIO, params=params, data=audio_file.read(), headers=headers)
    r_audio.raise_for_status()
    logging.info("Upload to Google Drive successful.")
    audio_file.seek(0) # Reset pointer again for the next upload

    # --- 2. Transcribe the Audio using Gemini ---
    logging.info("Starting transcription with Gemini...")
    
    # Upload the audio file to the Gemini API
    logging.info(f"Uploading {audio_filename} to Gemini for transcription...")
    
    # Create an in-memory binary stream to pass to the API
    audio_file_in_memory = io.BytesIO(audio_file.read())
    
    uploaded_audio = genai.upload_file(
        path=audio_file_in_memory,
        display_name=audio_filename,
        mime_type=audio_file.mimetype
    )
    logging.info("Audio uploaded to Gemini successfully.")

    # Prompt Gemini to transcribe the audio
    prompt = "Please transcribe the following audio recording. Provide only the text from the audio in english ONLY ENGLISH."
    response = gemini_model.generate_content([prompt, uploaded_audio])

    # Clean up the uploaded file on Google's servers
    genai.delete_file(uploaded_audio.name)

    if not response.text:
        logging.error("Gemini returned an empty transcription.")
        return jsonify({"error": "Transcription failed: The AI returned no text."}), 500
    
    transcription_text = response.text

    logging.info("Transcription complete:")
    logging.info(f"[{transcription_text}]")

    # --- 3. Analyze Text with Gemini ---
    logging.info("Sending to Gemini for analysis...")

    schema_info = TEMPLATE_SCHEMAS.get(template_name, TEMPLATE_SCHEMAS["template3"])
    chosen_schema = schema_info["schema"]

    prompt = f"""
    You are a data extraction assistant. Analyze the following transcription text.
    The user was dictating entries for the "{schema_info['name']}" template.
    Your task is to extract all the billing entries from the text and return them as a valid JSON list of objects.
    The required JSON format for EACH entry in the list is:
    {chosen_schema}
    - Be smart and map natural language (e.g., "flat B 502") to the correct JSON fields (e.g., "Tower No": "B", "Flat No": "502").
    - If a value is not mentioned, use "unknown" or 0.
    Transcription Text:
    "{transcription_text}"
    Return ONLY the JSON list. Do not include any other text or markdown.
    """

    response = gemini_model.generate_content(prompt)
    extracted_json_text = response.text.strip().replace("```json", "").replace("```", "")

    logging.info("Gemini analysis complete:")
    logging.info(f"{extracted_json_text}")

    # --- 4. Log Data to Google Sheet ---
    logging.info("Logging data to Google Sheet...")

    log_data = {
        "template": f"{template_name} ({schema_info['name']})",
        "transcription": transcription_text,
        "extractedData": extracted_json_text
    }

    r_log = requests.post(APPS_SCRIPT_URL_LOG, data=json.dumps(log_data), headers={'Content-Type': 'application/json'})
    r_log.raise_for_status()
    logging.info("Logging successful.")

    # --- 5. NEW: Create the Excel File ---
    logging.info("Creating Excel file...")

    # Load the JSON data
    try:
      data = json.loads(extracted_json_text)
      if not isinstance(data, list): # Ensure it's a list
        raise ValueError("Gemini did not return a list.")
    except Exception as e:
      logging.error(f"Error parsing JSON from Gemini: {e}")
      logging.error(f"Gemini output was: {extracted_json_text}")
      # Handle error - maybe create an empty Excel or one with the error
      return jsonify({"error": f"Could not understand AI output: {e}"}), 500

    # Create a new Excel workbook in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Data - {template_name}"

    # --- This is a complex part ---
    # We need to handle different templates.
    # This is a simple way that handles *most* cases.
    # A more robust solution would have different functions for each template.

    if not data:
      # Handle empty data
      ws['A1'] = "No data extracted."
      logging.warning("No data extracted from Gemini output.")
    else:
      # Get all headers from the first row of data
      # This handles different templates automatically
      headers = list(data[0].keys())
      logging.info(f"Excel headers: {headers}")

      # Write headers to the first row
      for col_idx, header in enumerate(headers, 1):
          ws[f"{{get_column_letter(col_idx)}}1"] = header

      # Write the data rows
      for row_idx, row_data in enumerate(data, 2): # Start from row 2
          for col_idx, header in enumerate(headers, 1):
              cell_value = row_data.get(header)

              # If the cell value is a list (like our 'expenses'),
              # just stringify it to put it in the cell.
              if isinstance(cell_value, list):
                  cell_value = json.dumps(cell_value)

              ws[f"{{get_column_letter(col_idx)}}{{row_idx}}"] = cell_value
      logging.info(f"Wrote {len(data)} rows to Excel file.")

    # Save the workbook to a in-memory file
    excel_in_memory = io.BytesIO()
    wb.save(excel_in_memory)
    excel_in_memory.seek(0) # Go to the start of the file

    logging.info("Excel file created successfully.")

    # --- 6. NEW: Send the Excel File to the User ---
    logging.info("Sending Excel file to the user.")
    return send_file(
        excel_in_memory,
        download_name='billing_report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True
    )
    # --- END OF ALL NEW LOGIC ---

  except Exception as e:
    logging.error(f"Error processing audio: {e}", exc_info=True)
    return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
  port = int(os.environ.get("PORT", 7860)) 
  # Run the app in production mode, accessible on the network
  app.run(debug=False, host='0.0.0.0', port=port)
